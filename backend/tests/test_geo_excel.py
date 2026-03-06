"""Tests for GEO Excel workbook generator."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.geo.excel_generator import generate_geo_workbook


def _complete_experiment():
    return {
        "id": 1,
        "name": "scRNA-seq of human PBMCs",
        "description": "Single-cell RNA-seq analysis.",
        "hypothesis": "Treatment X affects T-cell differentiation.",
        "owner_user_name": "Jane Doe",
        "samples": [
            {"organism": "Homo sapiens", "tissue_type": "blood"},
        ],
    }


def _complete_sample(sid: int = 101, name: str = "PBMC_001"):
    return {
        "id": sid,
        "sample_id_external": name,
        "organism": "Homo sapiens",
        "molecule_type": "total RNA",
        "tissue_type": "blood",
        "treatment_condition": "Vehicle control",
        "library_prep_method": "10x Chromium 3' v3.1",
        "library_layout": "paired",
        "prep_notes": "Standard protocol",
        "batch": {"instrument_model": "Illumina NovaSeq 6000"},
    }


def _complete_pipeline():
    return {
        "id": 10,
        "pipeline_name": "nf-core/scrnaseq",
        "pipeline_version": "2.7.0",
        "reference_genome": "GRCh38",
        "alignment_algorithm": "STARsolo",
    }


def _complete_files():
    return {
        "raw_files": [
            {"filename": "R1.fastq.gz", "md5_checksum": "abc123", "gcs_uri": "gs://b/R1.fastq.gz"},
        ],
        "processed_files": [
            {"filename": "matrix.h5", "md5_checksum": "def456", "gcs_uri": "gs://b/matrix.h5"},
        ],
        "raw_filenames": "R1.fastq.gz",
        "processed_filenames": "matrix.h5",
        "processed_gcs_uris": "gs://b/matrix.h5",
    }


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def test_workbook_has_three_sheets():
    wb_bytes = generate_geo_workbook(
        _complete_experiment(), [_complete_sample()], _complete_pipeline(), _complete_files()
    )
    wb = _load(wb_bytes)
    assert wb.sheetnames == ["SERIES", "SAMPLES", "PROTOCOLS"]


def test_series_sheet_headers_and_data():
    wb_bytes = generate_geo_workbook(
        _complete_experiment(), [_complete_sample()], _complete_pipeline(), _complete_files()
    )
    wb = _load(wb_bytes)
    ws = wb["SERIES"]

    # Row 1 is headers
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Series_title" in headers
    assert "Series_summary" in headers

    # Row 2 is data
    title_col = headers.index("Series_title") + 1
    assert ws.cell(row=2, column=title_col).value == "scRNA-seq of human PBMCs"


def test_samples_sheet_row_count():
    samples = [_complete_sample(101, "S001"), _complete_sample(102, "S002"), _complete_sample(103, "S003")]
    wb_bytes = generate_geo_workbook(_complete_experiment(), samples, _complete_pipeline(), _complete_files())
    wb = _load(wb_bytes)
    ws = wb["SAMPLES"]

    # 1 header row + 3 data rows
    assert ws.max_row == 4


def test_placeholder_for_missing_required():
    sample = _complete_sample()
    sample["organism"] = None  # Required field

    wb_bytes = generate_geo_workbook(_complete_experiment(), [sample], _complete_pipeline(), _complete_files())
    wb = _load(wb_bytes)
    ws = wb["SAMPLES"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    organism_col = headers.index("organism") + 1
    assert ws.cell(row=2, column=organism_col).value == "[REQUIRED - please fill in]"


def test_derived_library_strategy():
    wb_bytes = generate_geo_workbook(
        _complete_experiment(), [_complete_sample()], _complete_pipeline(), _complete_files()
    )
    wb = _load(wb_bytes)
    ws = wb["SAMPLES"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    strategy_col = headers.index("library_strategy") + 1
    assert ws.cell(row=2, column=strategy_col).value == "RNA-Seq"


def test_derived_library_source():
    wb_bytes = generate_geo_workbook(
        _complete_experiment(), [_complete_sample()], _complete_pipeline(), _complete_files()
    )
    wb = _load(wb_bytes)
    ws = wb["SAMPLES"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    source_col = headers.index("library_source") + 1
    assert ws.cell(row=2, column=source_col).value == "TRANSCRIPTOMIC"


def test_protocols_sheet_genome_build():
    wb_bytes = generate_geo_workbook(
        _complete_experiment(), [_complete_sample()], _complete_pipeline(), _complete_files()
    )
    wb = _load(wb_bytes)
    ws = wb["PROTOCOLS"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    genome_col = headers.index("Genome_build") + 1
    assert ws.cell(row=2, column=genome_col).value == "GRCh38"


def test_bold_headers():
    wb_bytes = generate_geo_workbook(
        _complete_experiment(), [_complete_sample()], _complete_pipeline(), _complete_files()
    )
    wb = _load(wb_bytes)

    for sheet_name in ["SERIES", "SAMPLES", "PROTOCOLS"]:
        ws = wb[sheet_name]
        for col in range(1, ws.max_column + 1):
            assert ws.cell(row=1, column=col).font.bold is True

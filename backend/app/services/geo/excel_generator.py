"""GEO Excel workbook generator — creates pre-filled GEO submission template."""

import json
import logging
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger("bioaf.geo.excel")

TEMPLATE_PATH = Path(__file__).parent / "geo_template_definition.json"
PLACEHOLDER = "[REQUIRED - please fill in]"


def _load_template() -> dict:
    with open(TEMPLATE_PATH) as f:
        return json.load(f)


def _derive_value(
    derivation: str,
    derivation_rules: dict[str, str] | None,
    source_value: str | None,
    experiment_data: dict,
    sample_data: dict | None,
    pipeline_data: dict | None,
) -> str | None:
    """Compute derived GEO field values. Reuses logic from validation module."""
    from app.services.geo.validation import _derive_value as _validate_derive

    return _validate_derive(derivation, derivation_rules, source_value, experiment_data, sample_data, pipeline_data)


def _get_source_value(
    bioaf_source: str,
    experiment_data: dict,
    sample_data: dict | None,
    pipeline_data: dict | None,
    files_data: dict | None,
) -> Any:
    """Resolve a bioAF source path to a value."""
    from app.services.geo.validation import _get_source_value as _validate_get

    return _validate_get(bioaf_source, experiment_data, sample_data, pipeline_data, files_data)


def _resolve_field_value(
    field_def: dict,
    experiment_data: dict,
    sample_data: dict | None,
    pipeline_data: dict | None,
    files_data: dict | None,
) -> str:
    """Resolve a field definition to its string value, or placeholder if missing."""
    derivation = field_def.get("derivation")

    if derivation:
        source_val = None
        if derivation == "library_strategy" and sample_data:
            source_val = sample_data.get("library_prep_method")
        elif derivation == "library_source" and sample_data:
            source_val = sample_data.get("molecule_type")
        elif derivation == "library_selection" and sample_data:
            source_val = sample_data.get("library_prep_method")

        value = _derive_value(
            derivation,
            field_def.get("derivation_rules"),
            source_val,
            experiment_data,
            sample_data,
            pipeline_data,
        )
    elif field_def["bioaf_source"] == "derived":
        value = None
    else:
        value = _get_source_value(
            field_def["bioaf_source"],
            experiment_data,
            sample_data,
            pipeline_data,
            files_data,
        )

    if value is None or (isinstance(value, str) and not value.strip()):
        if field_def.get("required", False):
            return PLACEHOLDER
        return ""

    return str(value)


def _auto_size_columns(ws: Any) -> None:
    """Set column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        # Cap at 60 chars, minimum 12
        adjusted = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[col_letter].width = adjusted


def generate_geo_workbook(
    experiment_data: dict,
    samples_data: list[dict],
    pipeline_data: dict | None,
    files_data: dict | None,
) -> bytes:
    """Generate GEO submission Excel workbook.

    Returns the workbook as bytes (xlsx format).
    """
    template = _load_template()
    wb = Workbook()

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")

    # --- SERIES sheet ---
    ws_series = wb.active
    ws_series.title = "SERIES"

    series_fields = template["sections"]["SERIES"]["fields"]
    for col_idx, field_def in enumerate(series_fields, 1):
        cell = ws_series.cell(row=1, column=col_idx, value=field_def["geo_column"])
        cell.font = header_font
        cell.alignment = header_alignment

    for col_idx, field_def in enumerate(series_fields, 1):
        value = _resolve_field_value(field_def, experiment_data, None, pipeline_data, files_data)
        ws_series.cell(row=2, column=col_idx, value=value)

    _auto_size_columns(ws_series)

    # --- SAMPLES sheet ---
    ws_samples = wb.create_sheet("SAMPLES")

    sample_fields = template["sections"]["SAMPLES"]["fields"]
    for col_idx, field_def in enumerate(sample_fields, 1):
        cell = ws_samples.cell(row=1, column=col_idx, value=field_def["geo_column"])
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, sample in enumerate(samples_data, 2):
        for col_idx, field_def in enumerate(sample_fields, 1):
            value = _resolve_field_value(field_def, experiment_data, sample, pipeline_data, files_data)
            ws_samples.cell(row=row_idx, column=col_idx, value=value)

    _auto_size_columns(ws_samples)

    # --- PROTOCOLS sheet ---
    ws_protocols = wb.create_sheet("PROTOCOLS")

    protocol_fields = template["sections"]["PROTOCOLS"]["fields"]
    first_sample = samples_data[0] if samples_data else None

    for col_idx, field_def in enumerate(protocol_fields, 1):
        cell = ws_protocols.cell(row=1, column=col_idx, value=field_def["geo_column"])
        cell.font = header_font
        cell.alignment = header_alignment

    for col_idx, field_def in enumerate(protocol_fields, 1):
        value = _resolve_field_value(field_def, experiment_data, first_sample, pipeline_data, files_data)
        ws_protocols.cell(row=2, column=col_idx, value=value)

    _auto_size_columns(ws_protocols)

    # Write to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
